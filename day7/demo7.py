# 导入 os 模块，用于处理文件路径和操作系统相关功能
import os

# 导入 random 模块，用于随机选择元素
import random

# 从 openpyxl 库中导入 load_workbook 函数，用于加载 Excel 文件
from openpyxl import load_workbook


# 拼接文件路径，将 "file" 目录和 "员工表.xlsx" 文件名组合成完整路径
file_path = os.path.join("file", "员工表.xlsx")

# 检查指定路径的文件是否存在
if not os.path.exists(file_path):
    # 若文件不存在，打印程序出错信息
    print("程序运行出错")
    # 明确告知用户文件不存在
    print("文件不存在")
else:
    # 2.2 读取 Excel 文件内容，并将内容存储到列表中
    # 初始化一个空列表，用于存储从 Excel 文件中读取的数据
    user_list = []

    # 加载指定路径的 Excel 文件
    wb = load_workbook(file_path)
    # 获取 Excel 文件的第一个工作表
    sheet = wb.worksheets[0]
    # 从第二行开始迭代工作表的每一行，读取 Excel 文件内容
    for row in sheet.iter_rows(min_row=2):
        # 获取当前行的第二个单元格（索引为 1）
        cell = row[1]
        # 将当前单元格的值添加到用户列表中
        user_list.append(cell.value)
        # 以下代码为注释示例，可将每行数据存储为字典

while True:
    text = input("请按任意键开始抽奖")

    if text == "q":
        break

    # 检查用户列表是否为空
    if not user_list:
        # 若用户列表为空，打印程序出错信息
        print("程序运行出错")
        # 明确告知用户列表为空

    # 从用户列表中随机选择一个元素
    user = random.choice(user_list)
    # 打印随机选择的元素
    messages = "恭喜{},{}中奖了！".format(user, "一等奖")
    print(messages)

    # 提出抽奖抽到的人
    user_list.remove(user)
