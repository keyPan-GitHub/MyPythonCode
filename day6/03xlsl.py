
from openpyxl import workbook


wb = workbook.Workbook()

sheet = wb.worksheets[0]
cell = sheet.cell(1,1).value = "开始"
cell = sheet.cell(1,2).value = "开始"
cell = sheet.cell(1,3).value = "开始"

wb.save("news.xlsx")