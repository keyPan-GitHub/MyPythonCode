import requests
from openpyxl import workbook

wb = workbook.Workbook()
sheet = wb.worksheets[0]

res = requests.get(
    "https://www.10086.cn/support/selfservice/help/sh/5010801_4073_8801.json"
)
data_dict = res.json()

data_list = data_dict["cData"]["list"]
row_index = 1
for item in data_list:
    question = item["question"]
    up_time = item["up_time"]
    id = item["_orderId"]
    # print(up_time,id,question)
    cell = sheet.cell(row_index, 1).value = up_time
    cell = sheet.cell(row_index, 2).value = id
    cell = sheet.cell(row_index, 3).value = question
    row_index += 1

wb.save("10086.xlsx")
